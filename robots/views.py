from django.http import JsonResponse
from django.views import View
from django.core.exceptions import ValidationError
from django.utils.dateparse import parse_datetime
from .models import Robot
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
import datetime
from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from robots.models import Robot


class RobotCreateView(View):
    @method_decorator(csrf_exempt, name='dispatch')
    def post(self, request, *args, **kwargs):
        try:
            import json
            data = json.loads(request.body)
            
            
            model = data.get('model')
            version = data.get('version')
            created = data.get('created')
            
            if not model or not version or not created:
                return JsonResponse({"error": "Missing required fields."}, status=400)
            
            
            created_date = parse_datetime(created)
            if not created_date:
                return JsonResponse({"error": "Invalid date format. Use ISO 8601."}, status=400)
            
            
            robot = Robot.objects.create(
                model=model,
                version=version,
                created=created_date,
                serial=f"{model}-{version}"
            )
            
            return JsonResponse({"message": "Robot created successfully.", "robot_id": robot.id}, status=201)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format."}, status=400)
        except ValidationError as e:
            return JsonResponse({"error": str(e)}, status=400)
        except Exception as e:
            return JsonResponse({"error": "An unexpected error occurred.", "details": str(e)}, status=500)
        


class WeeklyReportView(View):
    def get(self, request, *args, **kwargs):
        one_week_ago = datetime.datetime.now() - datetime.timedelta(days=7)

        robots = Robot.objects.filter(created__gte=one_week_ago)

        data = {}
        for robot in robots:
            if robot.model not in data:
                data[robot.model] = {}
            if robot.version not in data[robot.model]:
                data[robot.model][robot.version] = 0
            data[robot.model][robot.version] += 1

        wb = Workbook()
        if not data:
            ws = wb.active
            ws.title = "No Data"
            ws.cell(row=1, column=1, value="Нет данных за последнюю неделю")
        else:
            for model, versions in data.items():
                ws = wb.create_sheet(title=model)

                headers = ["Модель", "Версия", "Количество за неделю"]
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_num, value=header)

                for row_num, (version, count) in enumerate(versions.items(), start=2):
                    ws.cell(row=row_num, column=1, value=model)
                    ws.cell(row=row_num, column=2, value=version)
                    ws.cell(row=row_num, column=3, value=count)

                for col_num in range(1, 4):
                    column_width = max(len(str(cell.value)) for cell in ws[get_column_letter(col_num)])
                    ws.column_dimensions[get_column_letter(col_num)].width = column_width

            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=weekly_report.xlsx'
        wb.save(response)
        return response



